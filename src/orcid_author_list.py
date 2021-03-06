#!/usr/bin/env python
"""
Generating an author list based on a list of ORCIDs.

TODO:
    * Ordering authors correctly.
    * Matching affiliations based on disambiguated organization identifiers.

"""

import rtfunicode
import argparse as arg
import pdb
import openpyxl as xl

from itertools import chain
import csv

from Author import Author
from CanonicalAffiliations import CanonicalAffiliations


def __main__():

    parser = arg.ArgumentParser()
    parser.add_argument('--input', action='store', dest='input_file',
                        help='File containing one ORCID per line.')
    parser.add_argument('--comparison-view', action='store_true', dest='comparison_view',
                        help="""Output authors and affiliations in a mode condusive \
                        to harmonizing affiliations.""")
    parser.add_argument('--out', action='store', dest='output_file',
                        help='Output RTF file to write authors list to.')
    args = parser.parse_args()

    # load canonical affiliations
    affiliations_checker = CanonicalAffiliations()

    # load authors from orcid db
    orcid_contributions = list()
    fconn = csv.reader(open(args.input_file), delimiter='\t')
    next(fconn, None)  # skip header
    for line in fconn:
        orcid_contributions.append((line[1].strip(), [x.strip() for x in line[4].split(',')]))

    authors = {x[0]: Author(orcid=x[0], contributions=x[1], affiliations_checker=affiliations_checker) for x in orcid_contributions}
    # sort authors by contribution
    sorted_orcids = [x.orcid for x in sorted(authors.values())]

    # index unque affiliations based on author order in the orginal .csv file
    affiliations_index = dict()  # institution: printed index
    index_affiliations = dict()  # printed index: institution
    affiliations_seen_count = dict()
    counter = 1  # keep track of what order institutions appear in
    for orcid in sorted_orcids:
        if authors[orcid].affiliations:
            for affiliation in authors[orcid].affiliations:
                if affiliations_index.get(affiliation):
                    affiliations_seen_count[affiliation] += 1
                else:
                    affiliations_index[affiliation] = counter
                    index_affiliations[counter] = affiliation
                    counter += 1
                    affiliations_seen_count[affiliation] = 1
        else:
            continue

    # print author list: name and affilliation reference
    authors_with_affiliations = [authors[orcid]
                                 for orcid in sorted_orcids
                                 if authors[orcid].affiliations]
    # pdb.set_trace()
    author_affiliations = [author.affiliations
                           for author in authors_with_affiliations]

    if args.comparison_view:
        workbook = xl.Workbook(args.output_file)
        author_sheet = workbook.create_sheet('authors')

        sorted_affiliations = [index_affiliations[i + 1] for i in range(len(affiliations_index))]

        sorted_affiliations = sorted([x for x in sorted_affiliations], key=lambda x: x.institution_name.lower())

        affiliation_sheet = workbook.create_sheet('affiliations')
        header = ['author', 'affiliation_indices']
        author_sheet.append(header)

        formatted_affiliations_indices = [str(sorted([affiliations_index[aff] for aff in affs])).lstrip('[').rstrip(']') for affs in author_affiliations]
        encoded_authors = [str(author).encode('utf8') for author in authors_with_affiliations]
        for author, index in zip(encoded_authors, formatted_affiliations_indices):
            author_sheet.append((author, index))

        header = ['affiliation_index', 'affiliation_appearance_count', 'department', 'name', 'city', 'region', 'country', 'postal_code', 'disambiguated_id', 'source']
        affiliation_sheet.append(header)

        encoded_affiliations = (x.list_fields_utf() for x in sorted_affiliations)
        encoded_affiliation_index = (str(affiliations_index[i]).encode('utf8') for i in sorted_affiliations)
        encoded_affiliation_counts = (str(affiliations_seen_count[x]).encode('utf8') for x in sorted_affiliations)

        for index, affiliation, count in zip(encoded_affiliation_index, encoded_affiliations, encoded_affiliation_counts):
            affiliation_sheet.append(list(chain([index], [count], affiliation)))

        affiliation_changes_sheet = workbook.create_sheet('affiliations_changes')
        affiliation_changes_sheet.append(
            ("affiliation_type", "department", "institution_name", "city",
                "region", "country", "postal_code", "disambiguated_id",
                "disambiguation_source")
        )
        for observed, canonical in affiliations_checker._corrected_affiliations.items():
            affiliation_changes_sheet.append(['observed'] + observed.list_fields_utf())
            affiliation_changes_sheet.append(['canonical'] + canonical.list_fields_utf())

        workbook.save(args.output_file)
    else:
        with open(args.output_file, 'wb') as outFile:
            # rtf "header"
            out_bytes = b'{\\rtf1 \\utf-8 '
            formatted_affiliations_indices = ['{\\super ' + str(sorted([affiliations_index[aff]
                                              for aff in affs])).strip('[').strip(']') + '}' for affs in author_affiliations]
            encoded_affiliations_indicies = [x.encode('utf8') for x in formatted_affiliations_indices]

            encoded_authors = [str(author).encode('rtfunicode') for author in authors_with_affiliations]
            final_authors = (b''.join(x) for x in zip(encoded_authors, encoded_affiliations_indicies))
            out_bytes += b''.join((b'{' + author + b'\par}' for author in final_authors))

            sorted_affiliations = [index_affiliations[i + 1] for i in range(len(affiliations_index))]

            encoded_affiliations = [str(x).encode('rtfunicode') for x in sorted_affiliations]
            encoded_affiliation_index = ['{{\\super {0}}}'.format(affiliations_index[index_affiliations[i + 1]]).encode('utf8')
                                         for i in range(len(affiliations_index))]
            final_institutions = (b''.join(x) for x in zip(encoded_affiliation_index, encoded_affiliations))
            out_bytes += b''.join((b'{' + institution + b'\par}' for institution in final_institutions))

            # end rtf file
            out_bytes += b'}'

            outFile.write(out_bytes)


if __name__ == '__main__':
    __main__()
